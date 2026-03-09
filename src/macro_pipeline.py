"""Pipeline for downloading and persisting macroeconomic time series from FRED."""

from __future__ import annotations

import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import pandas as pd
from tqdm import tqdm

from src import config
from src.fred_client import fetch_series_observations
from src.utils import ensure_directory

logger = logging.getLogger(__name__)


def _empty_long_dataframe() -> pd.DataFrame:
    return pd.DataFrame(columns=["series_id", "series_name", "date", "value"])


def _observations_payload_to_frame(series_id: str, payload: dict) -> pd.DataFrame:
    observations = payload.get("observations") or []
    if not observations:
        return pd.DataFrame(
            {
                "series_id": pd.Series(dtype="string"),
                "series_name": pd.Series(dtype="string"),
                "date": pd.Series(dtype="datetime64[ns]"),
                "value": pd.Series(dtype="float64"),
            }
        )

    df = pd.DataFrame.from_records(observations)
    if not {"date", "value"}.issubset(df.columns):
        return _empty_long_dataframe()

    df = df[["date", "value"]].copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    df["series_id"] = series_id
    df["series_name"] = config.FRED_DEFAULT_SERIES.get(series_id, "")
    df = df.dropna(subset=["date"])
    return df


def build_macro_timeseries_long_dataframe(
    series_ids: list[str] | None = None,
    observation_start: str | None = None,
    observation_end: str | None = None,
) -> pd.DataFrame:
    """Build a long-format macro DataFrame: one row per (series_id, date)."""
    if not config.FRED_API_KEY:
        logger.warning(
            "FRED_API_KEY is not set; skipping macro download. Add it to `.env` as FRED_API_KEY=..."
        )
        return _empty_long_dataframe()

    selected_series_ids = series_ids or list(config.FRED_DEFAULT_SERIES.keys())
    start = observation_start or config.FRED_OBSERVATION_START
    end = (
        observation_end
        if observation_end is not None
        else config.EFFECTIVE_REFERENCE_DATE_STR
    )
    end = end or None

    frames: list[pd.DataFrame] = []
    max_workers = config.FRED_FETCH_MAX_WORKERS
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_series = {
            executor.submit(
                fetch_series_observations,
                series_id,
                observation_start=start,
                observation_end=end,
            ): series_id
            for series_id in selected_series_ids
        }
        for future in tqdm(
            as_completed(future_to_series),
            total=len(future_to_series),
            desc="Fetching FRED macro series",
            unit="series",
        ):
            series_id = future_to_series[future]
            try:
                payload = future.result()
            except Exception:
                logger.exception("Failed to fetch FRED series_id=%s", series_id)
                continue
            frames.append(_observations_payload_to_frame(series_id, payload))

    if not frames:
        return _empty_long_dataframe()

    df = pd.concat(frames, ignore_index=True)
    df = df.reindex(columns=["series_id", "series_name", "date", "value"])
    df = df.sort_values(["series_id", "date"], kind="stable").reset_index(drop=True)
    return df


def build_macro_timeseries_wide_dataframe(long_df: pd.DataFrame) -> pd.DataFrame:
    """Pivot long macro series into a wide DataFrame: one row per date, one col per series_id."""
    if long_df.empty:
        return pd.DataFrame()

    wide_df = long_df.pivot_table(
        index="date",
        columns="series_id",
        values="value",
        aggfunc="last",
    ).sort_index()
    wide_df.columns.name = None
    return wide_df


def build_macro_timeseries_quarterly_wide_dataframe(
    wide_df: pd.DataFrame,
    *,
    agg: str = "last",
) -> pd.DataFrame:
    """Resample wide macro series to quarter-end frequency (calendar quarters).

    Parameters
    ----------
    agg:
        Aggregation within each quarter. Supported: "last", "mean".
    """
    if wide_df.empty:
        return pd.DataFrame()

    df = wide_df.copy()
    if "date" in df.columns and not isinstance(df.index, pd.DatetimeIndex):
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df = df.dropna(subset=["date"]).set_index("date")

    if not isinstance(df.index, pd.DatetimeIndex):
        raise TypeError("wide_df must have a DatetimeIndex (or a 'date' column)")

    df = df.sort_index()
    resampler = df.resample("QE")
    if agg == "last":
        quarterly_df = resampler.last()
    elif agg == "mean":
        quarterly_df = resampler.mean()
    else:
        raise ValueError("agg must be one of: 'last', 'mean'")
    quarterly_df.columns.name = None
    return quarterly_df


def build_macro_timeseries_quarterly_delta_dataframe(
    quarterly_wide_df: pd.DataFrame,
) -> pd.DataFrame:
    """Compute percent change of a quarterly wide macro dataset."""
    if quarterly_wide_df.empty:
        return pd.DataFrame()
    prev = quarterly_wide_df.shift(1)
    delta_df = (quarterly_wide_df - prev) / prev * 100.0
    delta_df = delta_df.mask(prev == 0)
    delta_df.columns.name = None
    return delta_df


def _persist_wide_excel(
    out_path: Path,
    wide_df: pd.DataFrame,
    *,
    sheet_name: str,
) -> None:
    if wide_df.empty:
        return

    ensure_directory(out_path.parent)
    wide_out_df = wide_df.reset_index()
    if "date" in wide_out_df.columns:
        wide_out_df["date"] = pd.to_datetime(wide_out_df["date"], errors="coerce")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        wide_out_df.to_excel(writer, index=False, sheet_name=sheet_name)
        if "date" in wide_out_df.columns:
            from openpyxl.utils import get_column_letter

            date_col_idx = wide_out_df.columns.get_loc("date") + 1
            date_col_letter = get_column_letter(date_col_idx)
            ws = writer.book[sheet_name]
            for row in range(2, ws.max_row + 1):
                ws[f"{date_col_letter}{row}"].number_format = "yyyy-mm-dd"


def build_and_persist_macro_timeseries(
    series_ids: list[str] | None = None,
    observation_start: str | None = None,
    observation_end: str | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Download macro series from FRED and persist the wide Excel output."""
    long_df = build_macro_timeseries_long_dataframe(
        series_ids=series_ids,
        observation_start=observation_start,
        observation_end=observation_end,
    )
    wide_df = build_macro_timeseries_wide_dataframe(long_df)

    if wide_df.empty:
        return long_df, wide_df

    _persist_wide_excel(
        config.PROCESSED_MACRO_WIDE_PATH,
        wide_df,
        sheet_name="macro",
    )
    logger.info("Saved macro wide to %s", config.PROCESSED_MACRO_WIDE_PATH)

    quarterly_wide_df = build_macro_timeseries_quarterly_wide_dataframe(
        wide_df, agg="last"
    )
    if not quarterly_wide_df.empty:
        _persist_wide_excel(
            config.PROCESSED_MACRO_QUARTERLY_WIDE_PATH,
            quarterly_wide_df,
            sheet_name="macro_quarterly",
        )
        logger.info(
            "Saved macro quarterly wide to %s",
            config.PROCESSED_MACRO_QUARTERLY_WIDE_PATH,
        )

        quarterly_delta_df = build_macro_timeseries_quarterly_delta_dataframe(
            quarterly_wide_df
        )
        if not quarterly_delta_df.empty:
            _persist_wide_excel(
                config.PROCESSED_MACRO_QUARTERLY_DELTA_PATH,
                quarterly_delta_df,
                sheet_name="macro_quarterly_delta",
            )
            logger.info(
                "Saved macro quarterly delta to %s",
                config.PROCESSED_MACRO_QUARTERLY_DELTA_PATH,
            )

    quarterly_wide_mean_df = build_macro_timeseries_quarterly_wide_dataframe(
        wide_df, agg="mean"
    )
    if not quarterly_wide_mean_df.empty:
        _persist_wide_excel(
            config.PROCESSED_MACRO_QUARTERLY_WIDE_MEAN_PATH,
            quarterly_wide_mean_df,
            sheet_name="macro_quarterly_mean",
        )
        logger.info(
            "Saved macro quarterly wide mean to %s",
            config.PROCESSED_MACRO_QUARTERLY_WIDE_MEAN_PATH,
        )

        quarterly_delta_mean_df = build_macro_timeseries_quarterly_delta_dataframe(
            quarterly_wide_mean_df
        )
        if not quarterly_delta_mean_df.empty:
            _persist_wide_excel(
                config.PROCESSED_MACRO_QUARTERLY_DELTA_MEAN_PATH,
                quarterly_delta_mean_df,
                sheet_name="macro_quarterly_delta_mean",
            )
            logger.info(
                "Saved macro quarterly delta mean to %s",
                config.PROCESSED_MACRO_QUARTERLY_DELTA_MEAN_PATH,
            )

    return long_df, wide_df
